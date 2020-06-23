'''
    BEFORE RUNNING THE CODE...

    1. Before being able to run this code for the first time,
    make sure to install the openpyxl package by running the
    following command line in the terminal:

    pip install --user -U openpyxl==2.6.2

    2. Convert the hempBenchmarks pdf to an excel file using pdftoexcel.com

'''

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import date
from openpyxl.chart.axis import DateAxis

# input hempBenchmarksFile as 'fileName.xlsx' i.e. 'HempBenchmarksFebruary2020.xlsx'
# input monthAndYear as 'fullMonthName xxxx' i.e. 'September 2020'
# input prevMonthFile as 'fileName.xlsx' i.e. 'OctHempPricing.xlsx'
def dataFocus(hempBenchmarksFile, monthAndYear, prevMonthFile): # i.e. 'dataFocus('HempBenchmarksApril2020.xlsx', 'September 2020', 'NovHempPricing.xlsx')

    # open both excel files
    hb = openpyxl.load_workbook(hempBenchmarksFile, data_only = True)
    prevReport = openpyxl.load_workbook(prevMonthFile)
    newBook = openpyxl.Workbook()
    newBook.save(filename = 'HempPricing.xlsx')
    openpyxl.load_workbook('HempPricing.xlsx')

    # save sheet 1 as activeSheet
    activeSheet = hb['Sheet1']

    # edit new workbook sheet names
    newSheet1 = newBook['Sheet']
    newSheet1.title = monthAndYear[0:3] + ' ' + monthAndYear[-4:] + ' Data'
    newBook.create_sheet('Sheet2')
    newSheet2 = newBook['Sheet2']
    newSheet2.title = monthAndYear[0:3] + ' ' + monthAndYear[-4:] + ' Low Price Graph'
    newBook.create_sheet('Sheet3')
    newSheet3 = newBook['Sheet3']
    newSheet3.title = monthAndYear[0:3] + ' ' + monthAndYear[-4:] + ' High Price Graph'
    newBook.create_sheet('Sheet4')
    newSheet4 = newBook['Sheet4']
    newSheet4.title = 'Updated Line Graph'

    # edit prevMonthFile sheet names
    prevSheet4 = prevReport[prevReport.sheetnames[3]]
    newSheet4.column_dimensions['A'].width = 20

    # Writing to cells in newSheet1
    newSheet1['A1'] = 'U.S. Region Products'
    newSheet1['A1'].font = Font(bold = True)
    newSheet1['B1'] = 'Assessed Price'
    newSheet1['B1'].font = Font(bold = True)
    newSheet1['C1'] = 'Low'
    newSheet1['C1'].font = Font(bold = True)
    newSheet1['D1'] = 'High'
    newSheet1['D1'].font = Font(bold = True)
    newSheet1.column_dimensions['A'].width = 28
    newSheet1.column_dimensions['B'].width = 13

    # Writing to cells in newSheet2
    newSheet2['A1'] = 'U.S. Region Products'
    newSheet2['A1'].font = Font(bold = True)
    newSheet2['B1'] = 'Low'
    newSheet2['B1'].font = Font(bold = True)
    newSheet2['C1'] = 'Assessed Price'
    newSheet2['C1'].font = Font(bold = True)
    newSheet2['D1'] = 'High'
    newSheet2['D1'].font = Font(bold = True)
    newSheet2.column_dimensions['A'].width = 28
    newSheet2.column_dimensions['C'].width = 13

    # Writing to cells in newSheet3
    newSheet3['A1'] = 'U.S. Region Products'
    newSheet3['A1'].font = Font(bold = True)
    newSheet3['B1'] = 'Low'
    newSheet3['B1'].font = Font(bold = True)
    newSheet3['C1'] = 'Assessed Price'
    newSheet3['C1'].font = Font(bold = True)
    newSheet3['D1'] = 'High'
    newSheet3['D1'].font = Font(bold = True)
    newSheet3.column_dimensions['A'].width = 28
    newSheet3.column_dimensions['C'].width = 13

    # Writing to cells in newSheet4
    newSheet4.column_dimensions['A'].width = 11
    newSheet4.column_dimensions['B'].width = 14
    newSheet4.column_dimensions['C'].width = 14
    newSheet4.column_dimensions['D'].width = 20
    newSheet4.column_dimensions['E'].width = 18

    # Find the range of rows that the data is contained in (minRow to maxRow)
    for i in range(1, 45+1):
        if activeSheet.cell(row = i, column = 1).value == 'CBD Biomass (Aggregate)':
            minRow = i
        if activeSheet.cell(row = i, column = 1).value == 'CBG Isolate':
            maxRow = i

    # Populate data into newSheet1 (all data)
    lowIterNum = 0
    highIterNum = 0
    y = 2
    lowRow = 2
    highRow = 2
    for i in range(minRow,maxRow+1): # cycles through each row with data
        if activeSheet.cell(row = i, column = 4).value != None: # removes empty rows
            if activeSheet.cell(row = i, column = 4).value != 'Assessed Price': # removes title row
                # print(activeSheet.cell(row = i, column = 4).value)
                if activeSheet.cell(row = i, column = 4).value < 50: # for all rows cheaper than $50
                    lowIterNum += 1
                    lowProduct = activeSheet.cell(row = i, column = 1).value
                    lowPriceAssessed = activeSheet.cell(row = i, column = 4).value
                    lowPriceLow = activeSheet.cell(row = i, column = 5).value
                    lowPriceHigh = activeSheet.cell(row = i, column = 6).value
                    newSheet1.cell(row = y, column = 1).value = lowProduct
                    newSheet1.cell(row = y, column = 2).value = lowPriceAssessed
                    newSheet1.cell(row = y, column = 3).value = lowPriceLow
                    newSheet1.cell(row = y, column = 4).value = lowPriceHigh
                    y+=1
                    newSheet2.cell(row = lowRow, column = 1).value = lowProduct
                    newSheet2.cell(row = lowRow, column = 2).value = lowPriceLow
                    newSheet2.cell(row = lowRow, column = 3).value = lowPriceAssessed
                    newSheet2.cell(row = lowRow, column = 4).value = lowPriceHigh
                    lowRow+=1
                else:
                    highIterNum += 1
                    highProduct = activeSheet.cell(row = i, column = 1).value
                    highPriceAssessed = activeSheet.cell(row = i, column = 4).value
                    highPriceLow = activeSheet.cell(row = i, column = 5).value
                    highPriceHigh = activeSheet.cell(row = i, column = 6).value
                    newSheet1.cell(row = y, column = 1).value = highProduct
                    newSheet1.cell(row = y, column = 2).value = highPriceAssessed
                    newSheet1.cell(row = y, column = 3).value = highPriceLow
                    newSheet1.cell(row = y, column = 4).value = highPriceHigh
                    y+=1
                    newSheet3.cell(row = highRow, column = 1).value = highProduct
                    newSheet3.cell(row = highRow, column = 2).value = highPriceLow
                    newSheet3.cell(row = highRow, column = 3).value = highPriceAssessed
                    newSheet3.cell(row = highRow, column = 4).value = highPriceHigh
                    highRow+=1

    # At this point newSheet1 is complete!
    # At this point newSheet2 and newSheet3 have the correct data, just need graphs!

    # SHEET 4 STUFF...
    
    # Copy data from prevMonthFile to newSheet4
    for x in range(1,9):
        for y in range(1,6):
            newSheet4.cell(row = x, column = y).value = prevSheet4.cell(row = y, column = x).value

    # Save the month being pushed out off to the side
    for y in range(1,6):
        newSheet4.cell(row = 15, column = y).value = newSheet4.cell(row = 2, column = y).value

    # Update the cells by moving them all back 1 month
    for x in range(1,6):
        for y in range(3,9):
            newSheet4.cell(row = y-1, column = x).value = newSheet4.cell(row = y, column = x).value

    # Put in the new month's data for all 4 benchmarks
    newSheet4.cell(row = 7, column = 1).value = ' ' + monthAndYear[0:3] + ' ' + monthAndYear[-4:]
    for i in range(1,25):
        if newSheet1.cell(row = i, column = 1).value == 'Industrial Seeds':
            indSeedRow = i
        if newSheet1.cell(row = i, column = 1).value == 'CBD Clones':
            cloneRow = i
        if newSheet1.cell(row = i, column = 1).value == 'CBD Seeds (Feminized)':
            femSeedRow = i
        if newSheet1.cell(row = i, column = 1).value == 'CBD Biomass (Aggregate)':
            agMassRow = i

    indSeedPrice = newSheet1.cell(row = indSeedRow, column = 2).value
    clonePrice = newSheet1.cell(row = cloneRow, column = 2).value
    femSeedPrice = newSheet1.cell(row = femSeedRow, column = 2).value
    agMassPrice = newSheet1.cell(row = agMassRow, column = 2).value
        
    newSheet4.cell(row = 7, column = 2).value = indSeedPrice
    newSheet4.cell(row = 7, column = 3).value = clonePrice
    newSheet4.cell(row = 7, column = 4).value = femSeedPrice
    newSheet4.cell(row = 7, column = 5).value = agMassPrice

    # Place holders for the forecast
    newSheet4.cell(row = 8, column = 1).value = 'Forecasted Month'
    newSheet4.cell(row = 8, column = 2).value = 2.5
    newSheet4.cell(row = 8, column = 3).value = 2.5
    newSheet4.cell(row = 8, column = 4).value = 2.5
    newSheet4.cell(row = 8, column = 5).value = 2.5

    # Now delete extra data that I dont want graphed in newSheet2
    for i in range(1,15):
        if newSheet2.cell(row = i, column = 1).value == 'CBD Biomass (0 - 25K pounds)':
            newSheet2.delete_rows(i)
        if newSheet2.cell(row = i, column = 1).value == 'CBD Biomass (25K - 100K pounds)':
            newSheet2.delete_rows(i)
        if newSheet2.cell(row = i, column = 1).value == 'CBD Biomass (100K - 1M pounds)':
            newSheet2.delete_rows(i)
        if newSheet2.cell(row = i, column = 1).value == 'CBD Biomass (1M+ pounds)':
            newSheet2.delete_rows(i)
        if newSheet2.cell(row = i, column = 1).value == 'CBG Seeds':
            newSheet2.delete_rows(i)
        if newSheet2.cell(row = i, column = 1).value == 'CBG Clones':
            newSheet2.delete_rows(i)

    # Now delete extra data that I dont want graphed in newSheet3
    for i in range(1,15):
        if newSheet3.cell(row = i, column = 1).value == 'Distillate - THC Free':
            newSheet3.delete_rows(i)
        if newSheet3.cell(row = i, column = 1).value == 'Distillate - Broad Spectrum':
            newSheet3.delete_rows(i)
    
    # At this point the data for ALL 4 SHEETS ARE COMPLETE
    # Now we need graphs for sheets 2, 3, and 4...

    # Making the graph for sheet2
    chart1 = BarChart()
    chart1.type = 'bar'
    chart1.style = 10
    chart1.y_axis.title = 'USD'

    data = Reference(newSheet2, min_col=2, min_row=1, max_row=6, max_col=4)
    cats = Reference(newSheet2, min_col=1, min_row=2, max_row=7)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    y = chart1.height
    x = chart1.width
    chart1.height = y * 1.25
    chart1.width = x * 1.25
    newSheet2.add_chart(chart1, 'C9')

    # Making the graph for sheet3
    chart2 = BarChart()
    chart2.type = 'bar'
    chart2.style = 10
    chart2.y_axis.title = 'USD'

    data = Reference(newSheet3, min_col=2, min_row=1, max_row=9, max_col=4)
    cats = Reference(newSheet3, min_col=1, min_row=2, max_row=10)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.shape = 4
    y2 = chart2.height
    x2 = chart2.width
    chart2.height = y2 * 1.25
    chart2.width = x2 * 1.25
    newSheet3.add_chart(chart2, 'C12')

    # At this point, SHEETS 1, 2, AND 3 ARE NOW COMPLETED! now for pesky sheet 4...

    # All we have left is the sheet 4 graph:
    lineChart = LineChart()
    lineChart.style = 12
    lineChart.y_axis.title = 'USD'
    lineChart.y_axis.crossAx = 500
    lineChart.x_axis = DateAxis(crossAx=100)
    lineChart.x_axis.number_format = 'd-mmm'
    lineChart.x_axis.majorTimeUnit = 'days'

    prices = Reference(newSheet4, min_col=2, min_row=1, max_col=5, max_row=8)
    lineChart.add_data(prices, titles_from_data=True)
    dates = Reference(newSheet4, min_col=1, min_row=2, max_row=8)
    lineChart.set_categories(dates)

    newSheet4.add_chart(lineChart, "G2")


    # Save the new file
    newBook.save(monthAndYear[0:3] + 'HempPricing.xlsx')

    # close all files
    hb.close()
    newBook.close()
    prevReport.close()
